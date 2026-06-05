"""
fetch_efsa.py
-------------
Fetches and ingests the EU One Health 2022-2023 AMR Zoonoses data
from Zenodo (doi: 10.5281/zenodo.14645440) into amr_sentinel.db.

Sources:
- Annex C: Indicator E. coli (pigs, cattle, broilers, turkeys, meat)
- Annex A.2: Salmonella in food-producing animals
"""
import os, sys, logging, requests, sqlite3
from pathlib import Path
from io import BytesIO
import openpyxl
from dotenv import load_dotenv
load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

DB_PATH   = os.getenv("DATABASE_PATH", "amr_sentinel.db")
DATA_DIR  = Path("data")
DATA_DIR.mkdir(exist_ok=True)

ZENODO_BASE = "https://zenodo.org/records/14645440/files"

FILES = {
    "ecoli":      f"{ZENODO_BASE}/Annex%20C_Indicator%20E.%20coli_EFSA-ECDC_EUSR_AMR_2022-2023.xlsm?download=1",
    "salmonella": f"{ZENODO_BASE}/Annex%20A.2_Salmonella_food_producing_animals_EFSA-ECDC_EUSR_AMR_2022-2023.xlsm?download=1",
}

# Map sheet names to source_type labels
ECOLI_SHEETS = {
    "T. 1. Pigs E. coli":           "animal_pig",
    "T. 3. Broilers E. coli":       "animal_broiler",
    "T. 5. Pig meat BCP E. coli":   "meat_pork",
    "T. 7. Broiler meat BCP E. coli": "meat_broiler",
}

SALMONELLA_SHEETS = {
    "Pigs Salmonella spp.":    ("Salmonella spp.", "animal_pig"),
    "Pigs S. Derby":           ("S. Derby",        "animal_pig"),
    "Pigs S. Typhimurium":     ("S. Typhimurium",  "animal_pig"),
    "Pigs S. monophasic":      ("S. Typhimurium (mono)", "animal_pig"),
    "Broilers Salmonella spp.":("Salmonella spp.", "animal_broiler"),
    "Broilers S. Infantis":    ("S. Infantis",     "animal_broiler"),
    "Broilers S. Kentucky":    ("S. Kentucky",     "animal_broiler"),
    "Broilers S. Enteritidis": ("S. Enteritidis",  "animal_broiler"),
}

ANTIBIOTICS = ["GEN","AMK","CHL","AMP","CTX","CAZ","MEM","TGC",
               "NAL","CIP","AZM","COL","SMX","TMP","TET"]

YEAR = 2023  # data covers 2022-2023 reporting period


def download_file(url: str, dest: Path) -> Path:
    if dest.exists():
        log.info("Using cached %s", dest)
        return dest
    log.info("Downloading %s ...", dest.name)
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    dest.write_bytes(r.content)
    log.info("Saved %s (%.1f KB)", dest.name, len(r.content)/1024)
    return dest


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS amr_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            country TEXT, year INTEGER, organism TEXT,
            antibiotic TEXT, pct_resistant REAL,
            total_isolates INTEGER, source TEXT, source_type TEXT
        )""")
    conn.commit()
    return conn


def parse_pct(val):
    """Convert percentage value to float, handling strings like '1.7'."""
    if val is None:
        return None
    try:
        return float(str(val).replace('%','').strip())
    except:
        return None


def ingest_ecoli(conn, path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    inserted = 0
    for sheet_name, source_type in ECOLI_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            log.warning("Sheet not found: %s", sheet_name)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        # Build antibiotic column index map
        ab_idx = {}
        for ab in ANTIBIOTICS:
            for i, h in enumerate(header):
                if h and ab in str(h).upper().replace(' ',''):
                    ab_idx[ab] = i
                    break
        n_idx = next((i for i, h in enumerate(header) if h == 'N'), None)

        for row in rows[1:]:
            country = row[0]
            if not country or str(country).strip() in ('', 'EU/EEA', 'Total'):
                continue
            n_isolates = int(row[n_idx]) if n_idx and row[n_idx] else 0
            for ab, idx in ab_idx.items():
                pct = parse_pct(row[idx])
                if pct is None:
                    continue
                conn.execute("""
                    INSERT INTO amr_data
                    (country, year, organism, antibiotic, pct_resistant,
                     total_isolates, source, source_type)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (str(country).strip(), YEAR, "E. coli", ab,
                      pct, n_isolates, "EFSA_ECDC_2023", source_type))
                inserted += 1
    conn.commit()
    log.info("E. coli: inserted %d rows", inserted)


def ingest_salmonella(conn, path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    inserted = 0
    for sheet_name, (organism, source_type) in SALMONELLA_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            log.warning("Sheet not found: %s", sheet_name)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        ab_idx = {}
        for ab in ANTIBIOTICS:
            for i, h in enumerate(header):
                if h and ab in str(h).upper().replace(' ',''):
                    ab_idx[ab] = i
                    break
        n_idx = next((i for i, h in enumerate(header) if h == 'N'), None)

        for row in rows[1:]:
            country = row[0]
            if not country or str(country).strip() in ('', 'EU/EEA', 'Total'):
                continue
            n_isolates = int(row[n_idx]) if n_idx and row[n_idx] else 0
            for ab, idx in ab_idx.items():
                pct = parse_pct(row[idx])
                if pct is None:
                    continue
                conn.execute("""
                    INSERT INTO amr_data
                    (country, year, organism, antibiotic, pct_resistant,
                     total_isolates, source, source_type)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (str(country).strip(), YEAR, organism, ab,
                      pct, n_isolates, "EFSA_ECDC_2023", source_type))
                inserted += 1
    conn.commit()
    log.info("Salmonella: inserted %d rows", inserted)


def main():
    conn = get_connection()

    # Remove old EFSA data before re-ingesting
    conn.execute("DELETE FROM amr_data WHERE source LIKE 'EFSA%'")
    conn.commit()
    log.info("Cleared old EFSA data")

    ecoli_path      = download_file(FILES["ecoli"],
                                    DATA_DIR / "efsa_ecoli_2023.xlsm")
    salmonella_path = download_file(FILES["salmonella"],
                                    DATA_DIR / "efsa_salmonella_2023.xlsm")

    ingest_ecoli(conn, ecoli_path)
    ingest_salmonella(conn, salmonella_path)

    total = conn.execute(
        "SELECT COUNT(*) FROM amr_data WHERE source='EFSA_ECDC_2023'"
    ).fetchone()[0]
    log.info("Total EFSA 2023 rows in DB: %d", total)
    conn.close()


if __name__ == "__main__":
    main()
